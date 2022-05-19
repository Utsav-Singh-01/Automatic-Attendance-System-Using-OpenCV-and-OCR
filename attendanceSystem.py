#Misc modules used 
from datetime import date, datetime
import time
import os

#Modules related to opencv
import cv2 as cv
import numpy as np

#Modules related to ocr
from PIL import Image
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\\Program Files\\Tesseract-OCR\\tesseract'

#Module related to excel
import openpyxl
from openpyxl import workbook,load_workbook


#This function displays all the sheets(subjects) in the selected workbook and then returns the selected worksheet
def displaySubjects_and_returnSelected(selectedWB):
    print('\nThe courses of this batch are :-')
    i=1
    for ws in selectedWB.sheetnames:
        print(f"{i}.{ws}")
        i+=1

    selectedSubjectNumber = int(input("Please Enter the number of the subject you want to select :"))
    selectedSubjectNumber -=1
    selectedSubjectName = selectedWB.sheetnames[selectedSubjectNumber]
    selected_WS = selectedWB[selectedSubjectName]

    return selected_WS , selectedSubjectName


#Displays the existing batches and returns the workbook and the batch name according to the user's choice
def displayBatches_and_returnSelected():
    
    workbook_names = []
    batch_names = []

    i=1
    print('\nThe list of batches are :-')
    for filename in os.listdir('attendanceData'):
        workbook_names.append(filename)
        filename = filename.split('_')
        branch = filename[0]
        batch = branch + " " + filename[1]
        batch_names.append(batch)
        print(f"{i}.{batch}")
        i= i+1


    batchNumber = input('\nPlease Enter the number of the batch you want to select :')
    batchNumber = int(batchNumber)
    batchNumber = batchNumber-1


    selectedBatchName = batch_names[batchNumber]
    selectedWB = load_workbook('attendanceData/'+workbook_names[batchNumber])


    return selectedWB,selectedBatchName


#Fucntion that shows the attendance of all the students of a specific batch
def showAttendance(selected_WS,batch,course):
    os.system('cls')
    print('\nATTENDANCE RECORD OF',batch)
    print('\nCourse :',course,'\n\n')

    gap = ' '*3
    heading = f"{'Roll No.':13s}{gap}{'Student Name':20s}{gap}{'Total Present':13s}{gap}{'Total Absent':12s}"
    print('='*(13+20+13+12+10))
    print(heading)
    print('='*(13+20+13+12+10))
    rowNumber = 1
    totalDates = 0
    for row in selected_WS.iter_rows():
        if(rowNumber==1):
            for cell in row:
                if cell.column>3:
                    if cell.value!= None:
                        totalDates +=1
                        #print(cell.value)

               
        if(rowNumber>3):
            studentNameBreak = row[2].value.split(' ')
            if(len(studentNameBreak)>1):
                studentName = studentNameBreak[0] +' '+ studentNameBreak[1]
            else:
                studentName = studentNameBreak[0]
            
            totalAbsent = totalDates-int(row[0].value)
            rowPrint = f"{row[1].value}{gap}{studentName:20s}{gap}{row[0].value:13d}{gap}{totalAbsent:12d}"
            print(rowPrint)

        rowNumber = rowNumber + 1

    input("\nPress Enter to go back to Main Menu")
    os.system('cls')        


#Makes the user select a date and returns the date
def dateSelect_and_returnSelected():
    print('\nPlease select the method to input the date from the following options :-\n')
    print('\n1.Enter Date Manually')
    print("2.Automatically Select Today's Date\n")

    opt = int(input("Please enter the number of the date input method :"))

    if opt == 1:
        date = input('Enter the date in which you want to mark attendance (Format : dd/mm/yyyy) : ')
        return date

    if opt == 2:
        date = datetime.today()
        date = date.strftime("%d/%m/%Y")
        return str(date)


#Function that marking the attendance in the excel sheet
def markingAttendanceInExcel(roll,attendance_ws,attendance_date,selected_wb,batch):
    #First we have to find / append the date column
    dateColumnNo = -1
    dateFound = False
    
    #Searching For the date column
    for cell in attendance_ws[1]:
        if cell.value == date:
            dateColumnNo = cell.column
            dateFound = True
            break
    
    #Appending the date column if date not found
    if dateFound == False:
        #First we find the length of the dates row 
        dateColumnNo = len(attendance_ws[1])+1
        
        #Then we add new date at the end of the row
        attendance_ws.cell(row = 1,column = dateColumnNo).value=attendance_date


    #Now we have to search for the row of the student according to the roll to mark present
    studentRowNo = -1
    studentFound = False

    for cell in attendance_ws['B']:
        if cell.value == roll:
            studentRowNo = cell.row
            studentFound = True
            break

    if studentFound == False:
        print(f'Student with roll : {roll} is not a part of this batch.')
        return

    #First we mark the attendance in sheet by writing 'P' in the specific cell 
    #print(studentRowNo,dateColumnNo)
    attendance_ws.cell(row = studentRowNo,column = dateColumnNo).value = 'P'
    print(f'Attendance of roll : {roll} is succesfull marked')

    #Now we increase the total attendance of that student by 1
    attendance_ws.cell(row = studentRowNo,column = 1).value +=1

    saveFileName = f"attendanceData/{batch.split(' ')[0]}_{batch.split(' ')[1]}_Attendance.xlsx"
    selected_wb.save(saveFileName)



#Opencv code that read the roll from the ID and calls the function to mark the attendance in excel for that roll 
def markAttendance(attendance_date,batch,attendance_ws,selected_wb):
    
    branch = batch.split(' ')[0]

    # Using the webcam
    idCapture = cv.VideoCapture(0,cv.CAP_DSHOW)

    while(True):
        # Getting the frames from cam one by one 
        ret,frame = idCapture.read()
        if ret:
            # Putting Text
            idText ='Show ID'
            dataText = 'Keep Roll in this box'
            markText = 'Press M : To mark Attendance'
            endText = 'Press Esc : To save and close'
            cv.putText(frame,idText,(250,80),cv.FONT_HERSHEY_COMPLEX,1,(255,0,0),2)
            cv.putText(frame,dataText,(150,140),cv.FONT_HERSHEY_SIMPLEX,1,(0,0,255),2)
            cv.putText(frame,markText,(10,400),cv.FONT_HERSHEY_SIMPLEX,1,(0,255,255),2)
            cv.putText(frame,endText,(10,450),cv.FONT_HERSHEY_SIMPLEX,1,(0,255,255),2)
            # Putting the roll drectangle 
            cv.rectangle(frame,(200,160),(450,360),color=(0,0,255),thickness=2)
            cv.imshow('cam',frame)

            key = cv.waitKey(1)
            # When Esc is pressed 
            if  key == 27:
                print('Closing Attendance System')
                break

            # When M is pressed 
            if key==ord('m'):
                # Getting the Required part from the image (Croping)
                required_part = frame[160:360,200:450]
                savedName = 'tempsave/req_part.jpg' 
                # Converting Image to grayscale 
                required_part = cv.cvtColor(required_part,cv.COLOR_BGR2GRAY)

                # Sharpning the image
                kernel = np.array([[0, -1,  0],[-1,  5, -1],[0, -1,  0]])
                required_part = cv.filter2D(src=required_part, ddepth=-1, kernel=kernel)

                cv.imwrite(savedName,required_part)
                
                #OCR section
                text = pytesseract.image_to_string(Image.open(savedName))
                # splitting the text into different lines 
                lines = text.split('\n')
                # Finding line that contains the roll 
                found = False
                roll = ''
                for line in lines:
                    if(line[:4]=='2K20' or line[:2]=='2K' or line[2:4]=='20'):
                        roll = line[-3:]
                        found = True
                        #Checking if the roll string is numeric
                        if(roll.isnumeric()):
                            roll = '2K20/'+branch+'/'+roll
                            print("Marking the attendance of Roll :",roll)

                        else:
                            found = False

                if (found == True):
                    markingAttendanceInExcel(roll,attendance_ws,attendance_date,selected_wb,batch)        

                elif(found == False):
                    print("Couldn't Read Roll Please Try Again!")

    idCapture.release()
    cv.destroyAllWindows()        


#MENU
option = -1

while option!=0:
    
    print('\n\t\tMAIN MENU')
    print('\n1.Mark Attendance')
    print('2.View Attendance')
    print('3.Close')


    option = input('\nEnter your option number :')
    if option.isnumeric():
        option = int(option)
    else:
        continue

    if option ==1:
        os.system('cls')
        print('\n\tMark Attendance')
        print('\nSelect Batch :-',end='')
        
        selectedBatchWB , selectedBatchName = displayBatches_and_returnSelected()

        selectedSubject_WS , selectedCourseName = displaySubjects_and_returnSelected(selectedBatchWB)

        selectedDate = dateSelect_and_returnSelected()

        os.system('cls')
        print(f"Marking Attendance of {selectedBatchName} for course : {selectedCourseName}")
        markAttendance(selectedDate,selectedBatchName,selectedSubject_WS,selectedBatchWB) #calling the mark attendance function



    if option == 2:
        os.system('cls')
        print('\n\tVIEW ATTENDANCE')
       
        #First we make the user select a batch
        selectedBatchWB , selectedBatchName = displayBatches_and_returnSelected()
         
        #Then we make the user select the course/subject 
        selectedSubject_WS , selectedCourseName = displaySubjects_and_returnSelected(selectedBatchWB)        
        showAttendance(selectedSubject_WS,selectedBatchName,selectedCourseName)

    if option == 3:
        break

print('Closing the Attendance System ....')
time.sleep(1.5)
os.system('cls')